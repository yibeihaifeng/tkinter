# -- coding: utf-8 --
# @info:
# @Author : liyahui
# @Time : 2023/6/2 上午8:33
# @File : performance_testing.py
# @Software: PyCharm

"""电池性能测试"""
import datetime
import configparser
import os.path
import sys
import threading
import time
import tkinter as tk
from tkinter import ttk
from tkinter import *
import tkinter.messagebox as msgbox
from LogFormat import LogFormat
from UPS3072Adapter import UPS3072Adapter
import openpyxl

log = LogFormat('performance_testing', "performance_testing.log")
logger = log.logger


class Performance(tk.Frame):
    def __init__(self, config_path, master=tk.Tk()):
        super().__init__(master)
        self.log_text = None
        self.treeview_scrollbar = None
        self.root = master
        self.exit_button = None
        self.tree_view = None
        self.clear_text_button = None
        self.entry_box = None
        self.background = "#8FBC8F"
        self.font = "微软雅黑"
        self.clear_button = None
        self.execute_button = None
        self.start_height = None
        self.start_width = None
        self.box_width = None
        self.filepath = None
        self.text_box = None
        self.adapter_object = UPS3072Adapter('./conf/UPS3072Adapter.conf')
        self.config_path = config_path
        self.init_conf()
        self.run()

    def init_conf(self):
        """加载配置文件，初始化配置信息"""
        config = configparser.ConfigParser()
        try:
            config.read(self.config_path, encoding='utf-8')
            self.start_width = config.getint('performance_testing', 'width')
            self.start_height = config.getint('performance_testing', 'height')
            self.box_width = config.getint('performance_testing', 'box_width')
            self.filepath = config.get('performance_testing', 'filepath')
        except Exception as e:
            error_msg = f'init config file error：{e}'
            logger.error(error_msg)
            return False, error_msg

    def run(self):
        """界面初始化"""
        self.root.title("Performance testing")
        self.root.geometry(f"{self.start_width}x{self.start_height}+100+50")
        # Create the menu
        menu_bar = tk.Menu(self.root)
        # Create the file menu
        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="退出", command=self.exit)

        # Create the help menu
        help_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="关于", command=self.about)

        # Add the menus to the menu bar
        menu_bar.add_cascade(label="文件", menu=file_menu)
        menu_bar.add_cascade(label="帮助", menu=help_menu)

        # Add the menu bar to the window
        self.root.config(menu=menu_bar)
        # 第一行
        label = tk.Label(self.root, text="SN:", font=(self.font, 12))

        self.entry_box = tk.Entry(self.root, width=self.box_width, highlightcolor='red', highlightthickness=1)
        self.execute_button = tk.Button(self.root, text="执行测试", command=self.execute, width=10,
                                        background=self.background, font=(self.font, 12))
        self.execute_button.bind('<Return>', lambda event=None: self.execute_button.invoke())
        self.clear_button = tk.Button(self.root, text="清空", command=self.clear_input, width=10,
                                      font=(self.font, 12), background=self.background)

        label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.entry_box.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.execute_button.grid(row=0, column=2, padx=5, pady=5)
        self.clear_button.grid(row=0, column=3, padx=5, pady=5)
        # 第二行
        result_label = tk.Label(self.root, text="测试结果", font=(self.font, 12))
        result_label.grid(row=1, column=0, padx=5, pady=5)
        # Create the result box
        # 第三行
        self.clear_text_button = tk.Button(self.root, text="全部清空", command=self.clear_text, width=10,
                                           background=self.background,
                                           font=(self.font, 12))
        self.exit_button = tk.Button(self.root, text="退出程序", command=self.exit, width=10,
                                     background=self.background,
                                     font=(self.font, 12))

        style = ttk.Style(self.root)
        style.configure('Treeview.Heading', font=('Helvetica', 12, 'bold'))
        columns_list = ["序号", "测试项目", "方法", "3.2", "3.3", "3.4", "结果判断", "描述"]
        self.tree_view = ttk.Treeview(self.root, columns=columns_list, height=20,
                                      show="headings")
        self.tree_view.grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky="w")  #
        # 设置样式
        self.tree_view.configure(style='Treeview')
        self.clear_text_button.grid(row=3, column=2, padx=5, pady=5)
        self.exit_button.grid(row=3, column=3, padx=5, pady=5)
        # 第四行
        self.log_text = tk.Text(self.root,height=10)
        self.log_text.grid(row=4, column=1, columnspan=3, padx=5, pady=5,sticky="nwse")
        # Start the main loop
        self.root.mainloop()

    def clear_input(self):
        """清空输入框"""
        self.entry_box.delete(0, "end")
        logger.info("清空输入框")

    def clear_text(self):
        """全部清空"""
        self.entry_box.delete(0, "end")
        self.log_text.delete('1.0',"end")
        for column in self.tree_view["columns"]:
            self.tree_view.heading(column, text="")
        self.tree_view.delete(*self.tree_view.get_children())
        logger.info("全部清空")

    def get_log(self):
        """获取日志信息"""
        q = self.adapter_object.log_q
        while True:
            if not q.empty():
                log_msg = q.get()
                date_info = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.log_text.insert(1.0, f"{date_info} {log_msg}" + "\r\n")
            else:
                pass
            time.sleep(1)

    def execute(self):
        """执行测试"""
        self.execute_button['state'] = 'disable'
        sn_text = self.entry_box.get().strip()  # 获取输入框中的文本，去掉两侧的换行符
        logger.info(f"设备:{sn_text},开始测试")
        th = threading.Thread(target=self.thread_execute, args=(sn_text,), daemon=True)
        th.start()
        get_log_th = threading.Thread(target=self.get_log, daemon=True)
        get_log_th.start()

    def thread_execute(self, sn_text):
        """线程执行测试"""
        time_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_file_str = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        filename = f"SN_{sn_text}_{time_file_str}.xlsx"
        log_msg = ""
        is_ok = False
        if not os.path.exists(self.filepath):
            log_msg = f"找不到文件路径：{self.filepath}"
        else:
            abs_filepath = os.path.join(self.filepath, filename)
            # abs_filepath = os.path.join(r"C:\Users\lyh\Desktop\工作文件\UPS3702\测试报告",
            #                             "data.xlsx")
            try:
                self.adapter_object.connect()
                is_ok, log_msg = self.adapter_object.read_hold_register(abs_filepath, time_str, sn_text)
                # is_ok = True
            except Exception as e:
                log_msg = f"执行测试出错:{e}"
                logger.info(log_msg)

            else:
                if is_ok:  # 测试成功
                    # 读取xlsx文件内容
                    rows = []
                    if os.path.exists(abs_filepath):
                        try:
                            workbook = openpyxl.load_workbook(abs_filepath)
                            worksheet = workbook.active
                        except Exception as e:
                            log_msg = f"打开测试报告出错:{e}"
                        else:
                            if worksheet.max_row <= 2 or worksheet.max_column <= 4:
                                log_msg = f"测试报告内容为空"
                            else:
                                for row in worksheet.iter_rows(values_only=True):
                                    rows.append(row)
                                # 设置列名
                                self.tree_view['columns'] = tuple(rows[1])
                                self.treeview_scrollbar = ttk.Scrollbar(self.root, orient='vertical',
                                                                        command=self.tree_view.yview)
                                self.treeview_scrollbar.grid(row=2, column=4, sticky=(tk.N, tk.S))
                                self.tree_view.configure(yscrollcommand=self.treeview_scrollbar.set)

                                # 设置 heading
                                for i, column in enumerate(rows[0]):
                                    self.tree_view.heading(i, text=column)
                                for row_index, row in enumerate(rows[1:]):
                                    if row_index == 0:
                                        table_head_list = list(row)
                                        table_head_list = [item.replace("\n", "") for item in table_head_list]
                                        row = tuple(table_head_list)
                                        self.tree_view.insert("", tk.END, values=row)
                                    else:
                                        self.tree_view.insert("", tk.END, values=row)
                                log_msg = f"测试完成！"
                    else:
                        log_msg = f"找不到报告：{abs_filepath}"

        logger.info(log_msg)
        msgbox.showinfo(title="提示", message=log_msg)
        self.execute_button['state'] = 'normal'

    def exit(self):
        """退出"""
        self.root.destroy()
        # sys.exit(0)# 终止线程
        logger.info(f"退出")

    @staticmethod
    def about(self):
        """返回版本号"""
        version = "1.0"
        show_msg = f"版本:{version}"
        msgbox.showinfo(title="提示", message=show_msg)
        logger.info(f"关于:{show_msg}")


if __name__ == "__main__":
    performance = Performance('./conf/performance_testing.conf')
    performance.mainloop()
